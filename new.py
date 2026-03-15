import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# -------------------------------
# 1. Read uploaded file
# -------------------------------
def read_input(uploaded_file):
    if uploaded_file is None:
        return None
    if uploaded_file.name.endswith(".csv"):
        df = pd.read_csv(uploaded_file)
    else:
        df = pd.read_excel(uploaded_file)
    return df

# -------------------------------
# 2. Preprocess transactions
# -------------------------------
def preprocess_data(df, pending=False):
    df.columns = df.iloc[5]
    df = df.iloc[6:].reset_index(drop=True)

    # Ensure Date column
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.date

    # Keep useful columns
    cols = ["Date", "Appears On Your Statement As", "Amount", "Category", "Description", "Card Member", "Account #"]
    df = df[cols].copy()
    df = df.rename(columns={
        "Appears On Your Statement As": "Location",
        "Description": "Descript"
    })

    if pending:
        df["Category"] = "PENDING"

    df["JOBS"] = " - "
    df["Notes"] = ""

    return df

# -------------------------------
# 3. Combine and group for Excel
# -------------------------------
def group_and_sort(df_combined):
    dfs, order = [], []
    for (member, acct), group in df_combined.groupby(["Card Member", "Account #"]):
        group = group.sort_values(by=["Date", "Amount"], ascending=[False, False]).reset_index(drop=True)
        group = group[["JOBS", "Date", "Location", "Amount", "Category", "Descript", "Notes"]]
        group.columns = ["JOBS", "Date", "Location", "Amount", "Category", "Description", "Notes"]
        dfs.append(group)
        order.append((member, acct))
    final = pd.concat(dfs, axis=1)
    return final, order

# -------------------------------
# 4. Style Excel
# -------------------------------
def style_excel(final, order):
    output = BytesIO()
    final.to_excel(output, index=False)
    output.seek(0)

    wb = load_workbook(output)
    ws = wb.active

    colors = ["FFCCCC", "CCFFCC", "CCCCFF", "FFFFCC", "FFCCFF", "CCFFFF", "E0E0E0"]
    col_offset = 1

    for idx, (member, acct) in enumerate(order):
        width = 7
        start_col, end_col = col_offset, col_offset + width - 1

        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        ws.cell(row=1, column=start_col).value = f"{member} {acct}"
        ws.cell(row=1, column=start_col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        fill = PatternFill(start_color=colors[idx % len(colors)], end_color=colors[idx % len(colors)], fill_type="solid")
        for col in range(start_col, end_col + 1):
            ws.cell(row=1, column=col).fill = fill
            ws.cell(row=1, column=col).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        amount_col = start_col + 3
        for row in range(3, ws.max_row + 1):
            cell = ws.cell(row=row, column=amount_col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = u'"$"#,##0.00'
                if cell.value > 300:
                    cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                elif cell.value > 75:
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True)

        for col in range(start_col, end_col + 1):
            if col != amount_col:
                for row in range(3, ws.max_row + 1):
                    ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

        col_offset += width

    final_buf = BytesIO()
    wb.save(final_buf)
    final_buf.seek(0)
    return final_buf

st.title("💳 Combined Expense Formatter")

# Upload both files
uploaded_posted = st.file_uploader("Upload Posted Transactions", type=["xlsx", "csv"])
uploaded_pending = st.file_uploader("Upload Pending Transactions", type=["xlsx", "csv"])

if uploaded_posted and uploaded_pending:
    df_posted = read_input(uploaded_posted)
    df_pending = read_input(uploaded_pending)

    if df_posted is not None and df_pending is not None:
        df_posted = preprocess_data(df_posted, pending=False)
        df_pending = preprocess_data(df_pending, pending=True)

        # Combine
        df_combined = pd.concat([df_posted, df_pending], ignore_index=True)

        # Group, sort
        final_df, order = group_and_sort(df_combined)

        # Style and produce Excel buffer
        excel_buf = style_excel(final_df, order)

        st.download_button(
            label="⬇️ Download Combined Excel",
            data=excel_buf,
            file_name="combined_expenses.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )