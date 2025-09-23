# %%
import pandas as pd 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill

# %%
data = pd.read_excel("activity (18).xlsx")

# %%
data.columns = data.iloc[5]
data = data.iloc[6:]


# %%
data

# %%
data = data.sort_values(by=["Card Member", "Category", "Amount"]).reset_index(drop=True)


# %%

# # Example: starting from your `data` DataFrame
# data = data.sort_values(by=["Card Member", "Category", "Amount"]).reset_index(drop=True)

# # Keep only the useful columns
# cols = ["Date", "Appears On Your Statement As", "Amount", "Category", "Description", "Card Member"]
# df = data[cols].copy()

# # Rename for readability
# df = df.rename(columns={
#     "Appears On Your Statement As": "Location",
#     "Description": "Descript"
# })

# # Create a "JOB" index like Excel rows
# df["JOBS"] = [f"JOB-{i+1}" for i in range(len(df))]

# # Split dataframe by Card Member
# dfs = []
# for member, group in df.groupby("Card Member"):
#     group = group.reset_index(drop=True)
#     group = group[["JOBS", "Date", "Location", "Amount", "Category", "Descript"]]
#     # rename columns with card member name prefix
#     group.columns = pd.MultiIndex.from_product([[f"{member}"], group.columns])
#     dfs.append(group)

# # Concatenate side by side
# final = pd.concat(dfs, axis=1)

# final.to_csv("expenses_by_cardmember.csv", index=False)


# %%
import pandas as pd

# -----------------------
# Step 1: Prepare the data
# -----------------------
data = data.sort_values(by=["Account #", "Card Member", "Category", "Amount"]).reset_index(drop=True)

cols = ["Date", "Appears On Your Statement As", "Amount", "Category", "Description", "Card Member", "Account #"]
df = data[cols].copy()
df = df.rename(columns={
    "Appears On Your Statement As": "Location",
    "Description": "Descript"
})

df["JOBS"] = ""

# -----------------------
# Step 2: Build per-cardholder groups
# -----------------------
dfs = []
order = []
for (member, acct), group in df.groupby(["Card Member", "Account #"]):
    group = group.reset_index(drop=True)
    group = group[["JOBS", "Date", "Location", "Amount", "Category", "Descript"]]
    group.columns = ["JOBS", "Date", "Location", "Amount", "Category", "Description"]
    dfs.append(group)
    order.append((member, acct))

# concat side by side
final = pd.concat(dfs, axis=1)

# -----------------------
# Step 3: Export to Excel
# -----------------------
output_file = "expenses_by_cardmember.xlsx"
final.to_excel(output_file, index=False)

# -----------------------
# Step 4: Post-process Excel
# -----------------------
wb = load_workbook(output_file)
ws = wb.active

# Define fill colors (rotate if more cardholders than colors)
colors = [
    "FFCCCC",  # light red
    "CCFFCC",  # light green
    "CCCCFF",  # light blue
    "FFFFCC",  # light yellow
    "FFCCFF",  # light pink
    "CCFFFF",  # light cyan
    "E0E0E0",  # grey
]

col_offset = 1
for idx, (member, acct) in enumerate(order):
    width = 6  # JOBS..Description
    start_col = col_offset
    end_col = col_offset + width - 1

    # Merge top row
    ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
    ws.cell(row=1, column=start_col).value = f"{member} {acct}"
    ws.cell(row=1, column=start_col).alignment = Alignment(horizontal="center", vertical="center")

    # Apply fill color to header row (row 1 + column headers row 2)
    fill = PatternFill(start_color=colors[idx % len(colors)], end_color=colors[idx % len(colors)], fill_type="solid")
    for col in range(start_col, end_col + 1):
        ws.cell(row=1, column=col).fill = fill
        ws.cell(row=2, column=col).fill = fill

    # Format Amount column (4th column inside block) as currency
    amount_col = start_col + 3  # JOBS(1), Date(2), Location(3), Amount(4)
    for row in range(3, ws.max_row + 1):  # data rows start at row 3
        cell = ws.cell(row=row, column=amount_col)
        if isinstance(cell.value, (int, float)):
            cell.number_format = u'"$"#,##0.00'
    for col in range(start_col, end_col + 1):
        for row in range(3, ws.max_row + 1):
            ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")
    col_offset += width

wb.save(output_file)



